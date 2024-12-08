import os
import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from tkinter import messagebox
import logging
from logging.handlers import RotatingFileHandler

# Setup logging
def setup_logging(log_dir="logs"):
    """Menyiapkan logging dengan rotasi file dan log level."""
    if not os.path.exists(log_dir):
        os.makedirs(log_dir)

    log_file = os.path.join(log_dir, "app.log")
    handler = RotatingFileHandler(log_file, maxBytes=5*1024*1024, backupCount=3)
    handler.setLevel(logging.DEBUG)  # Mengatur level log secara global
    formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    handler.setFormatter(formatter)
    
    logger = logging.getLogger()
    logger.setLevel(logging.DEBUG)
    logger.addHandler(handler)

def log_message(terminal, message, index=None, total=None):
    """Menambahkan log ke terminal, dengan informasi urutan data."""
    if index is not None and total is not None:
        message = f"Data {index}/{total}: {message}"  # Menambahkan nomor urut
    terminal.insert("end", message + "\n")
    terminal.see("end")
    logging.info(message)  # Simpan log ke file

def get_unique_filename(folder, filename):
    """Menghasilkan nama file unik jika sudah ada file dengan nama yang sama."""
    base, ext = os.path.splitext(filename)
    counter = 1
    new_filename = filename
    while os.path.exists(os.path.join(folder, new_filename)):
        new_filename = f"{base}({counter}){ext}"
        counter += 1
    return new_filename

def check_internet_connection():
    """Memeriksa koneksi internet."""
    try:
        response = requests.get("https://www.google.com", timeout=5)
        return response.status_code == 200
    except requests.exceptions.RequestException:
        return False

def process_data(input_path, output_folder, progress_bar, progress_label, terminal):
    """Memproses data dari input ke output dan menambahkan border ke Excel.""" 
    filename = "output.xlsx"
    unique_filename = get_unique_filename(output_folder, filename)
    output_file = os.path.join(output_folder, unique_filename)

    # Membaca file input
    with open(input_path, 'r') as file:
        input_data_list = [line.strip() for line in file.readlines()]
    tables_list = []

    # Mengatur progress bar
    progress_bar["maximum"] = len(input_data_list)

    # Proses data
    for index, input_data in enumerate(input_data_list, start=1):
        url = f'https://mrtg.iconpln.co.id/check/zabbix?s={input_data}'
        
        log_message(terminal, f"Meminta URL: {url}", index, len(input_data_list))

        try:
            response = requests.get(url)
            response.raise_for_status()  # Memeriksa apakah status code bukan 4xx/5xx
            tables = pd.read_html(url)
            if tables:
                table = tables[0]
                table.insert(0, 'Data', [''] * len(table))
                table.at[0, 'Data'] = input_data
                tables_list.append(table)
                log_message(terminal, f"Sukses: Data untuk {input_data}.", index, len(input_data_list))
            else:
                log_message(terminal, f"Tidak ada tabel untuk {input_data}.", index, len(input_data_list))
        except requests.exceptions.RequestException as e:
            log_message(terminal, f"Gagal untuk {input_data}: {e}", index, len(input_data_list))
        
        # Update progress
        progress_bar["value"] = index
        progress_label.config(text=f"Proses {index}/{len(input_data_list)} selesai")

    # Simpan ke file Excel
    if tables_list:
        df = pd.concat(tables_list, ignore_index=True)
        df.to_excel(output_file, index=False)

        # Load file Excel untuk memodifikasi
        wb = load_workbook(output_file)
        ws = wb.active

        # Menambahkan border untuk setiap sel
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
        
        # Menyesuaikan lebar kolom sesuai dengan panjang data
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Mendapatkan nama kolom
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)  # Menambahkan sedikit padding
            ws.column_dimensions[column].width = adjusted_width

        # Simpan perubahan
        wb.save(output_file)

        log_message(terminal, f"Data berhasil disimpan di {output_file}", index=None, total=None)
        messagebox.showinfo("Sukses", f"Proses selesai. File disimpan di {output_file}")
    else:
        log_message(terminal, "Tidak ada data yang diproses.", index=None, total=None)
        messagebox.showwarning("Peringatan", "Tidak ada data yang diproses.")
    
    return output_file
