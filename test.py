from tkinter import Tk, filedialog, messagebox, Button, Label, Entry, Text, Scrollbar, END
from tkinter.ttk import Progressbar
import os
import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from threading import Thread
import time

def log_message(message):
    """Fungsi untuk menambahkan log ke terminal."""
    terminal.insert(END, message + '\n')
    terminal.see(END)  # Scroll otomatis ke bawah

def get_unique_filename(folder, filename):
    """Menghasilkan nama file unik jika sudah ada file dengan nama yang sama."""
    base, ext = os.path.splitext(filename)
    counter = 1
    new_filename = filename
    while os.path.exists(os.path.join(folder, new_filename)):
        new_filename = f"{base}({counter}){ext}"
        counter += 1
    return new_filename

def process_data(input_path, output_folder):
    filename = "output.xlsx"
    unique_filename = get_unique_filename(output_folder, filename)
    output_file = os.path.join(output_folder, unique_filename)

    with open(input_path, 'r') as file:
        input_data_list = [line.strip() for line in file.readlines()]
    tables_list = []
    added_data = set()

    total_items = len(input_data_list)
    if total_items == 0:
        log_message("Tidak ada data untuk diproses.")
        return

    progress_bar['maximum'] = total_items

    start_time = time.time()
    for index, input_data in enumerate(input_data_list, start=1):
        if input_data in added_data:
            continue

        url = f'https://mrtg.iconpln.co.id/check/zabbix?s={input_data}'
        log_message(f"Meminta URL: {url}")

        try:
            response = requests.get(url)
            response.raise_for_status()

            if response.status_code == 200:
                try:
                    tables = pd.read_html(url)
                    if tables:
                        table = tables[0]
                        table.insert(0, 'Data', [''] * len(table))
                        table.at[0, 'Data'] = input_data
                        tables_list.append(table)
                        log_message(f"Sukses: Data untuk {input_data} diambil.")
                    else:
                        empty_row = pd.DataFrame([[input_data, None, None, None]], columns=['Data', 1, 2, 3])
                        tables_list.append(empty_row)
                        log_message(f"Data kosong untuk {input_data}.")
                except Exception as e:
                    empty_row = pd.DataFrame([[input_data, None, None, None]], columns=['Data', 1, 2, 3])
                    tables_list.append(empty_row)
                    log_message(f"Gagal membaca tabel untuk {input_data}: {e}")
        except requests.exceptions.RequestException as e:
            empty_row = pd.DataFrame([[input_data, None, None, None]], columns=['Data', 1, 2, 3])
            tables_list.append(empty_row)
            log_message(f"Error saat mengakses URL {url}: {e}")

        added_data.add(input_data)

        # Update progress bar
        progress_bar['value'] = index
        elapsed_time = time.time() - start_time
        estimated_total_time = elapsed_time / index * total_items
        remaining_time = estimated_total_time - elapsed_time
        progress_label.config(
            text=f"Proses {index}/{total_items} selesai. Estimasi sisa waktu: {int(remaining_time)} detik."
        )
        root.update_idletasks()

    df_with_no_spacing = pd.concat(tables_list, ignore_index=True)

    df_with_no_spacing.to_excel(output_file, index=False)

    wb = load_workbook(output_file)
    ws = wb.active

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    wb.save(output_file)
    log_message(f"Data berhasil disimpan ke {output_file}.")
    progress_label.config(text="Proses selesai.")
    return output_file

def select_input_file():
    input_path = filedialog.askopenfilename(filetypes=[("Text Files", "*.txt")])
    if input_path:
        input_entry.delete(0, "end")
        input_entry.insert(0, input_path)

def select_output_folder():
    output_folder = filedialog.askdirectory()
    if output_folder:
        output_entry.delete(0, "end")
        output_entry.insert(0, output_folder)

def run_processing():
    input_path = input_entry.get()
    output_folder = output_entry.get()
    if not os.path.exists(input_path):
        messagebox.showerror("Error", "File input tidak ditemukan.")
        return
    if not output_folder:
        messagebox.showerror("Error", "Tentukan folder output.")
        return

    def process():
        try:
            process_data(input_path, output_folder)
            messagebox.showinfo("Sukses", "Proses selesai.")
        except Exception as e:
            log_message(f"Error: {str(e)}")
            messagebox.showerror("Error", str(e))

    Thread(target=process).start()

# GUI Setup
root = Tk()
root.title("Data Processing App")

Label(root, text="File Input:").grid(row=0, column=0, padx=10, pady=10)
input_entry = Entry(root, width=50)
input_entry.grid(row=0, column=1, padx=10, pady=10)
Button(root, text="Browse", command=select_input_file).grid(row=0, column=2, padx=10, pady=10)

Label(root, text="Folder Output:").grid(row=1, column=0, padx=10, pady=10)
output_entry = Entry(root, width=50)
output_entry.grid(row=1, column=1, padx=10, pady=10)
Button(root, text="Browse", command=select_output_folder).grid(row=1, column=2, padx=10, pady=10)

Button(root, text="Run", command=run_processing).grid(row=2, column=0, columnspan=3, pady=10)

progress_bar = Progressbar(root, orient="horizontal", length=400, mode="determinate")
progress_bar.grid(row=3, column=0, columnspan=3, pady=10)

progress_label = Label(root, text="")
progress_label.grid(row=4, column=0, columnspan=3)

Label(root, text="Log Output:").grid(row=5, column=0, columnspan=3)

# Terminal Log
terminal = Text(root, height=15, width=70, state="normal")
terminal.grid(row=6, column=0, columnspan=3, padx=10, pady=10)

scrollbar = Scrollbar(root, command=terminal.yview)
scrollbar.grid(row=6, column=3, sticky="ns")
terminal.config(yscrollcommand=scrollbar.set)

root.mainloop()
